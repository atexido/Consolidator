import pandas as pd
import datetime
import openpyxl
import locale
import os
from openpyxl.styles.alignment import Alignment
from openpyxl.styles import Font, Color, colors
from openpyxl.styles import Font, PatternFill
from openpyxl.styles import numbers
from openpyxl import load_workbook

# Leer el archivo FINAL y tomar el valor de la primera fila de la columna TOTAL
df_final = pd.read_excel('M:/PAGOS/__Pagos/Pagos_Test/Pagos_Test/bin/Temp/MejorCredito/Sucursal/FINAL.xlsx')
total = df_final.loc[0, 'TOTAL']
num_filas_final = len(df_final)

df_finalsg = pd.read_excel('M:/PAGOS/__Pagos/Pagos_Test/Pagos_Test/bin/Temp/MejorCredito/Resultados/df_sin_gestion.xlsx')
totalsg = df_finalsg.loc[0, 'TOTAL']
num_filas_finalsg = len(df_finalsg)

# Leer el archivo viernes y abrir la hoja del mes concurrente
mes_actual = datetime.datetime.now().strftime('%B')
df_viernes = pd.read_excel('M:/PAGOS/MONITOR/Financieras/2023_financieras.xlsx', sheet_name=mes_actual)
cant_sem = len(df_viernes)

# Encontrar la columna con la fecha más cercana al día de hoy
fecha_hoy = datetime.datetime.now().date()
columna_fecha = min(df_viernes.columns, key=lambda fecha: abs(pd.to_datetime(fecha).date() - fecha_hoy))

# Escribir en la primera fila de esa columna el valor de la columna TOTAL del archivo FINAL

df_viernes.loc[0, columna_fecha] = total
df_viernes.loc[1, columna_fecha] = totalsg

meses = {'January': 'Enero', 'February': 'Febrero', 'March': 'Marzo', 'April': 'Abril',
         'May': 'Mayo', 'June': 'Junio', 'July': 'Julio', 'August': 'Agosto',
         'September': 'Septiembre', 'October': 'Octubre', 'November': 'Noviembre',
         'December': 'Diciembre'}

mes_actual_espanol = meses[mes_actual]



if os.path.isfile(f'{mes_actual_espanol}.xlsx'):
    # Leer el archivo existente y unirlo con el nuevo DataFrame
    with pd.ExcelWriter(f'{mes_actual_espanol}2.xlsx') as writer:
        
        num_cols = len(df_viernes.columns)
        print(num_cols)
        if(num_cols == 5):    
            new_cols = ['SEMANA{}'.format(i) for i in range(1, num_cols + 1)]
            df_viernes.columns = new_cols
            df_viernes['CLIENTE'] = ''
            df_viernes['TOTALIZADOR'] = ''
            df_viernes['CANT_PAGOS'] = ''
            df_viernes['OBJETIVO'] = ''
            df_viernes['%'] = ''
            df_viernes = df_viernes[['CLIENTE', 'SEMANA1', 'SEMANA2','SEMANA3', 'SEMANA4', 'SEMANA5', 'TOTALIZADOR', 'OBJETIVO', '%', 'CANT_PAGOS']]
            df_viernes.to_excel(writer, sheet_name=mes_actual, index=False)

        if(num_cols == 4):
            new_cols = ['SEMANA{}'.format(i) for i in range(1, num_cols + 1)]
            df_viernes.columns = new_cols
            df_viernes['CLIENTE'] = ''
            df_viernes['TOTALIZADOR'] = ''
            df_viernes['CANT_PAGOS'] = ''
            df_viernes['OBJETIVO'] = ''
            df_viernes['%'] = ''
            df_viernes = df_viernes[['CLIENTE', 'SEMANA1', 'SEMANA2','SEMANA3', 'SEMANA4', 'TOTALIZADOR', 'OBJETIVO', '%', 'CANT_PAGOS']]
            df_viernes.to_excel(writer, sheet_name=mes_actual, index=False)
    
    df1 = pd.read_excel(f'{mes_actual_espanol}.xlsx')
    df2 = pd.read_excel(f'{mes_actual_espanol}2.xlsx')

    for col in df1.columns:
        for i, value in enumerate(df2[col]):
            if pd.notnull(value):
               df1.at[i, col] = value

        # Guardar resultado en un nuevo archivo
        writer = pd.ExcelWriter(f'{mes_actual_espanol}.xlsx')
        df1.to_excel(writer, index=False)
        worksheet = writer.sheets['Sheet1']
        worksheet.set_column('A:J', 15)
        writer.save()
    
    os.remove(f'{mes_actual_espanol}2.xlsx')

else:
    # Guardar los cambios en el archivo viernes
    with pd.ExcelWriter(f'{mes_actual_espanol}.xlsx') as writer:
        df_viernes.to_excel(writer, sheet_name=mes_actual, index=False)
    df = pd.read_excel(f'{mes_actual_espanol}.xlsx')
    # renombrar columnas
    num_cols = len(df.columns)
    print(num_cols)
    if(num_cols == 5):
        new_cols = ['SEMANA{}'.format(i) for i in range(1, num_cols + 1)]
        df.columns = new_cols
        df['CLIENTE'] = ''
        df['TOTALIZADOR'] = ''
        df['CANT_PAGOS'] = ''
        df['OBJETIVO'] = ''
        df['%'] = ''
        df = df[['CLIENTE', 'SEMANA1', 'SEMANA2','SEMANA3', 'SEMANA4', 'SEMANA5', 'TOTALIZADOR', 'OBJETIVO', '%', 'CANT_PAGOS']]
        # guardar cambios en el archivo
        df.to_excel(f'{mes_actual_espanol}.xlsx', index=False)   

    if(num_cols == 4):
        new_cols = ['SEMANA{}'.format(i) for i in range(1, num_cols + 1)]
        df.columns = new_cols
        df['CLIENTE'] = ''
        df['TOTALIZADOR'] = ''
        df['CANT_PAGOS'] = ''
        df['OBJETIVO'] = ''
        df['%'] = ''
        df = df[['CLIENTE', 'SEMANA1', 'SEMANA2','SEMANA3', 'SEMANA4', 'TOTALIZADOR', 'OBJETIVO', '%', 'CANT_PAGOS']]
        # guardar cambios en el archivo
        df.to_excel(f'{mes_actual_espanol}.xlsx', index=False)   

# Lee el archivo xlsx
df_cereo = pd.read_excel(f'{mes_actual_espanol}.xlsx')
# Obtén las columnas que contienen la palabra SEMANA
semana_cols = [col for col in df_cereo.columns if 'SEMANA' in col]
# Recorre las columnas y verifica si alguna está vacía
for col in semana_cols:
    if df_cereo[col].isnull().values.any():
        # Si está vacía, coloca un 0 en la primera fila
        df_cereo.loc[df_cereo[col].isnull(), col] = 0
df_cereo.to_excel(f'{mes_actual_espanol}.xlsx', index=False)   


# Cargar el archivo xlsx
df = pd.read_excel(f'{mes_actual_espanol}.xlsx')
# Insertar valor en la primera fila de la columna "OBJETIVO"
df.loc[0, "OBJETIVO"] = "20000000"
# Guardar cambios en el archivo xlsx
df.to_excel(f'{mes_actual_espanol}.xlsx', index=False)
# Cargar el archivo xlsx


df = pd.read_excel(f'{mes_actual_espanol}.xlsx')

dfcli = df
dfcli.loc[0, "CLIENTE"] = "MEJORCREDITO"
dfcli.loc[1, "CLIENTE"] = "SIN GESTION"
dfcli.to_excel(f'{mes_actual_espanol}.xlsx', index=False)

dfsum = df
# Obtener las columnas que contienen la palabra SEMANA

#semanas = dfsum.filter(like='SEMANA')
#valor_maximo = semanas.max().max()

columnas_semana = [col for col in dfsum.columns if 'SEMANA' in col]

# encontrar el valor máximo de cada columna
maximos = []
for columna in columnas_semana:
    maximo = dfsum[columna].max()
    maximos.append(maximo)

# encontrar la columna con el mayor valor máximo
indice_maximo = maximos.index(max(maximos))
columna_maximo = columnas_semana[indice_maximo]

# sumar los valores de la columna con el mayor valor máximo
totalizador = dfsum[columna_maximo].sum()

# crear una nueva columna con el totalizador
dfsum['TOTALIZADOR'] = totalizador

#dfsum['TOTALIZADOR'] = valor_maximo
# Guardar el resultado en un nuevo archivo xlsx
dfsum.to_excel(f'{mes_actual_espanol}.xlsx', index=False)

dfporc1 = df
objetivo = df.loc[0, "OBJETIVO"]
dfporc1["%"] = dfporc1[columna_maximo] / objetivo * 100
dfporc1.to_excel(f'{mes_actual_espanol}.xlsx', index=False)

dfcant_pagos = df
dfcant_pagos.loc[0, "CANT_PAGOS"] = num_filas_final
dfcant_pagos.loc[1, "CANT_PAGOS"] = num_filas_finalsg
dfcant_pagos.to_excel(f'{mes_actual_espanol}.xlsx', index=False)


# Leer archivo xlsx
df_porcint = pd.read_excel(f'{mes_actual_espanol}.xlsx')
# Guardar los valores de la columna '%'
columna_porcentaje = df_porcint['%']
# Eliminar los decimales de los valores de la columna '%'
columna_porcentaje_sin_decimales = columna_porcentaje.astype(int)
# Reemplazar los valores originales de la columna '%' por los valores sin decimales
df_porcint['%'] = columna_porcentaje_sin_decimales
# Guardar el DataFrame actualizado en un nuevo archivo xlsx
df_porcint.to_excel(f'{mes_actual_espanol}.xlsx', index=False)
df3 = pd.read_excel(f'{mes_actual_espanol}.xlsx')

# ajustar el ancho de la columna A
writer = pd.ExcelWriter(f'{mes_actual_espanol}.xlsx', engine='xlsxwriter')
df3.to_excel(writer, sheet_name='Sheet1', index=False)
worksheet = writer.sheets['Sheet1']
worksheet.set_column('A:J', 15)
writer.save()

archivo = openpyxl.load_workbook(f'{mes_actual_espanol}.xlsx')
# Obtener la hoja de trabajo
hoja = archivo['Sheet1']
# Combinar las celdas de las dos filas de la columna H
hoja.merge_cells('H2:H3')
# Centrar el contenido de la celda combinada
celda = hoja['H2']
celda.alignment = Alignment(horizontal='center', vertical='center')
font = Font(color='FF0000')
celda.font = font
# Guardar los cambios en el archivo xlsx
archivo.save(f'{mes_actual_espanol}.xlsx')

archivo2 = openpyxl.load_workbook(f'{mes_actual_espanol}.xlsx')
# Obtener la hoja de trabajo
hoja2 = archivo2['Sheet1']
# Combinar las celdas de las dos filas de la columna H
hoja2.merge_cells('G2:G3')
# Centrar el contenido de la celda combinada
celda2 = hoja2['G2']
celda2.alignment = Alignment(horizontal='center', vertical='center')
font2 = Font(color='000000')
celda2.font = font2
# Guardar los cambios en el archivo xlsx
archivo2.save(f'{mes_actual_espanol}.xlsx')

# Cargar el archivo excel con pandas
dfcolor = pd.read_excel(f'{mes_actual_espanol}.xlsx')
# Obtener el nombre de la hoja del archivo
nombre_hoja = 'Sheet1'
# Crear una instancia de workbook de openpyxl
book = load_workbook(f'{mes_actual_espanol}.xlsx')
# Seleccionar la hoja del libro
sheet = book[nombre_hoja]
# Crear un objeto de estilo para las celdas
font = Font(color="FF0000")
# Recorrer cada celda de la columna deseada y aplicar el estilo de letra roja
for cell in sheet['H']:
    if isinstance(cell.value, str):
        cell.font = font
# Guardar los cambios en el archivo
book.save(f'{mes_actual_espanol}.xlsx')


df_paint = pd.ExcelFile(f'{mes_actual_espanol}.xlsx')
sem1 = df.loc[0, 'SEMANA1']
sem2 = df.loc[0, 'SEMANA2']
sem3 = df.loc[0, 'SEMANA3']
sem4 = df.loc[0, 'SEMANA4']
sem5 = df.loc[0, 'SEMANA5']

sem1T = ((sem1*100)/objetivo)
sem2T = ((sem2*100)/objetivo)
sem3T = ((sem3*100)/objetivo)
sem4T = ((sem4*100)/objetivo)
sem5T = ((sem5*100)/objetivo)


if sem1T < 18:
    # Leemos la hoja del archivo que deseamos
    hoja_excel = pd.read_excel(df_paint, sheet_name="Sheet1")
    # Creamos un estilo de fuente y fondo para aplicar a la columna SEMANA2
    fill_style = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    # Abrimos el archivo con openpyxl
    excel = load_workbook(filename=(f'{mes_actual_espanol}.xlsx'))
    # Seleccionamos la hoja donde queremos editar
    hoja_openpyxl = excel["Sheet1"]
    # Recorremos la columna SEMANA2 y le aplicamos el estilo creado anteriormente
    for celda in hoja_openpyxl["B"]:
        celda.fill = fill_style
    # Guardamos los cambios
    excel.save(f'{mes_actual_espanol}.xlsx')

if sem1T >= 18 and sem1T < 21:
# Leemos la hoja del archivo que deseamos
    hoja_excel = pd.read_excel(df_paint, sheet_name="Sheet1")
    # Creamos un estilo de fuente y fondo para aplicar a la columna SEMANA2
    fill_style = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    # Abrimos el archivo con openpyxl
    excel = load_workbook(filename=(f'{mes_actual_espanol}.xlsx'))
    # Seleccionamos la hoja donde queremos editar
    hoja_openpyxl = excel["Sheet1"]
    # Recorremos la columna SEMANA2 y le aplicamos el estilo creado anteriormente
    for celda in hoja_openpyxl["B"]:
        celda.fill = fill_style
    # Guardamos los cambios
    excel.save(f'{mes_actual_espanol}.xlsx')

if sem1T >= 21:
    # Leemos la hoja del archivo que deseamos
    hoja_excel = pd.read_excel(df_paint, sheet_name="Sheet1")
    # Creamos un estilo de fuente y fondo para aplicar a la columna SEMANA2
    fill_style = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    # Abrimos el archivo con openpyxl
    excel = load_workbook(filename=(f'{mes_actual_espanol}.xlsx'))
    # Seleccionamos la hoja donde queremos editar
    hoja_openpyxl = excel["Sheet1"]
    # Recorremos la columna SEMANA2 y le aplicamos el estilo creado anteriormente
    for celda in hoja_openpyxl["B"]:
        celda.fill = fill_style
    # Guardamos los cambios
    excel.save(f'{mes_actual_espanol}.xlsx')

if sem2T < 35:
    # Leemos la hoja del archivo que deseamos
    hoja_excel = pd.read_excel(df_paint, sheet_name="Sheet1")
    # Creamos un estilo de fuente y fondo para aplicar a la columna SEMANA2
    fill_style = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    # Abrimos el archivo con openpyxl
    excel = load_workbook(filename=(f'{mes_actual_espanol}.xlsx'))
    # Seleccionamos la hoja donde queremos editar
    hoja_openpyxl = excel["Sheet1"]
    # Recorremos la columna SEMANA2 y le aplicamos el estilo creado anteriormente
    for celda in hoja_openpyxl["C"]:
        celda.fill = fill_style
    # Guardamos los cambios
    excel.save(f'{mes_actual_espanol}.xlsx')

if sem2T >= 35 and sem1T < 40:
# Leemos la hoja del archivo que deseamos
    hoja_excel = pd.read_excel(df_paint, sheet_name="Sheet1")
    # Creamos un estilo de fuente y fondo para aplicar a la columna SEMANA2
    fill_style = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    # Abrimos el archivo con openpyxl
    excel = load_workbook(filename=(f'{mes_actual_espanol}.xlsx'))
    # Seleccionamos la hoja donde queremos editar
    hoja_openpyxl = excel["Sheet1"]
    # Recorremos la columna SEMANA2 y le aplicamos el estilo creado anteriormente
    for celda in hoja_openpyxl["C"]:
        celda.fill = fill_style
    # Guardamos los cambios
    excel.save(f'{mes_actual_espanol}.xlsx')

if sem3T >= 40:
    # Leemos la hoja del archivo que deseamos
    hoja_excel = pd.read_excel(df_paint, sheet_name="Sheet1")
    # Creamos un estilo de fuente y fondo para aplicar a la columna SEMANA2
    fill_style = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    # Abrimos el archivo con openpyxl
    excel = load_workbook(filename=(f'{mes_actual_espanol}.xlsx'))
    # Seleccionamos la hoja donde queremos editar
    hoja_openpyxl = excel["Sheet1"]
    # Recorremos la columna SEMANA2 y le aplicamos el estilo creado anteriormente
    for celda in hoja_openpyxl["C"]:
        celda.fill = fill_style
    # Guardamos los cambios
    excel.save(f'{mes_actual_espanol}.xlsx')

if sem3T < 55:
    # Leemos la hoja del archivo que deseamos
    hoja_excel = pd.read_excel(df_paint, sheet_name="Sheet1")
    # Creamos un estilo de fuente y fondo para aplicar a la columna SEMANA2
    fill_style = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    # Abrimos el archivo con openpyxl
    excel = load_workbook(filename=(f'{mes_actual_espanol}.xlsx'))
    # Seleccionamos la hoja donde queremos editar
    hoja_openpyxl = excel["Sheet1"]
    # Recorremos la columna SEMANA2 y le aplicamos el estilo creado anteriormente
    for celda in hoja_openpyxl["D"]:
        celda.fill = fill_style
    # Guardamos los cambios
    excel.save(f'{mes_actual_espanol}.xlsx')

if sem3T >= 55 and sem1T < 60:
# Leemos la hoja del archivo que deseamos
    hoja_excel = pd.read_excel(df_paint, sheet_name="Sheet1")
    # Creamos un estilo de fuente y fondo para aplicar a la columna SEMANA2
    fill_style = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    # Abrimos el archivo con openpyxl
    excel = load_workbook(filename=(f'{mes_actual_espanol}.xlsx'))
    # Seleccionamos la hoja donde queremos editar
    hoja_openpyxl = excel["Sheet1"]
    # Recorremos la columna SEMANA2 y le aplicamos el estilo creado anteriormente
    for celda in hoja_openpyxl["D"]:
        celda.fill = fill_style
    # Guardamos los cambios
    excel.save(f'{mes_actual_espanol}.xlsx')

if sem3T >= 60:
    # Leemos la hoja del archivo que deseamos
    hoja_excel = pd.read_excel(df_paint, sheet_name="Sheet1")
    # Creamos un estilo de fuente y fondo para aplicar a la columna SEMANA2
    fill_style = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    # Abrimos el archivo con openpyxl
    excel = load_workbook(filename=(f'{mes_actual_espanol}.xlsx'))
    # Seleccionamos la hoja donde queremos editar
    hoja_openpyxl = excel["Sheet1"]
    # Recorremos la columna SEMANA2 y le aplicamos el estilo creado anteriormente
    for celda in hoja_openpyxl["D"]:
        celda.fill = fill_style
    # Guardamos los cambios
    excel.save(f'{mes_actual_espanol}.xlsx')


if sem4T < 75:
    # Leemos la hoja del archivo que deseamos
    hoja_excel = pd.read_excel(df_paint, sheet_name="Sheet1")
    # Creamos un estilo de fuente y fondo para aplicar a la columna SEMANA2
    fill_style = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    # Abrimos el archivo con openpyxl
    excel = load_workbook(filename=(f'{mes_actual_espanol}.xlsx'))
    # Seleccionamos la hoja donde queremos editar
    hoja_openpyxl = excel["Sheet1"]
    # Recorremos la columna SEMANA2 y le aplicamos el estilo creado anteriormente
    for celda in hoja_openpyxl["E"]:
        celda.fill = fill_style
    # Guardamos los cambios
    excel.save(f'{mes_actual_espanol}.xlsx')

if sem4T >= 75 and sem1T < 80:
# Leemos la hoja del archivo que deseamos
    hoja_excel = pd.read_excel(df_paint, sheet_name="Sheet1")
    # Creamos un estilo de fuente y fondo para aplicar a la columna SEMANA2
    fill_style = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    # Abrimos el archivo con openpyxl
    excel = load_workbook(filename=(f'{mes_actual_espanol}.xlsx'))
    # Seleccionamos la hoja donde queremos editar
    hoja_openpyxl = excel["Sheet1"]
    # Recorremos la columna SEMANA2 y le aplicamos el estilo creado anteriormente
    for celda in hoja_openpyxl["E"]:
        celda.fill = fill_style
    # Guardamos los cambios
    excel.save(f'{mes_actual_espanol}.xlsx')

if sem4T >= 80:
    # Leemos la hoja del archivo que deseamos
    hoja_excel = pd.read_excel(df_paint, sheet_name="Sheet1")
    # Creamos un estilo de fuente y fondo para aplicar a la columna SEMANA2
    fill_style = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    # Abrimos el archivo con openpyxl
    excel = load_workbook(filename=(f'{mes_actual_espanol}.xlsx'))
    # Seleccionamos la hoja donde queremos editar
    hoja_openpyxl = excel["Sheet1"]
    # Recorremos la columna SEMANA2 y le aplicamos el estilo creado anteriormente
    for celda in hoja_openpyxl["E"]:
        celda.fill = fill_style
    # Guardamos los cambios
    excel.save(f'{mes_actual_espanol}.xlsx')


if sem5T < 80:
    hoja_excel = pd.read_excel(df_paint, sheet_name="Sheet1")
    fill_style = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    excel = load_workbook(filename=(f'{mes_actual_espanol}.xlsx'))
    hoja_openpyxl = excel["Sheet1"]
    for celda in hoja_openpyxl["F"]:
        celda.fill = fill_style
    excel.save(f'{mes_actual_espanol}.xlsx')

if sem5T >= 95 and sem1T < 100:
    hoja_excel = pd.read_excel(df_paint, sheet_name="Sheet1")
    fill_style = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    excel = load_workbook(filename=(f'{mes_actual_espanol}.xlsx'))
    hoja_openpyxl = excel["Sheet1"]
    for celda in hoja_openpyxl["F"]:
        celda.fill = fill_style
    excel.save(f'{mes_actual_espanol}.xlsx')

if sem5T >= 100:
    hoja_excel = pd.read_excel(df_paint, sheet_name="Sheet1")
    fill_style = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    excel = load_workbook(filename=(f'{mes_actual_espanol}.xlsx'))
    hoja_openpyxl = excel["Sheet1"]
    for celda in hoja_openpyxl["F"]:
        celda.fill = fill_style
    excel.save(f'{mes_actual_espanol}.xlsx')


df_moneda = pd.read_excel(f'{mes_actual_espanol}.xlsx')

# Dar formato de moneda a la columna SEMANA2
wb = openpyxl.load_workbook(f'{mes_actual_espanol}.xlsx')
ws = wb.active

for cell in ws['B'][1:]:
    cell.number_format = '$#,##0.00'
    wb.save(f'{mes_actual_espanol}.xlsx')

wb = openpyxl.load_workbook(f'{mes_actual_espanol}.xlsx')
ws = wb.active

for cell in ws['C'][1:]:
    cell.number_format = '$#,##0.00'
    wb.save(f'{mes_actual_espanol}.xlsx')

wb = openpyxl.load_workbook(f'{mes_actual_espanol}.xlsx')
ws = wb.active

for cell in ws['D'][1:]:
    cell.number_format = '$#,##0.00'
    wb.save(f'{mes_actual_espanol}.xlsx')

wb = openpyxl.load_workbook(f'{mes_actual_espanol}.xlsx')
ws = wb.active

for cell in ws['E'][1:]:
    cell.number_format = '$#,##0.00'
    wb.save(f'{mes_actual_espanol}.xlsx')

wb = openpyxl.load_workbook(f'{mes_actual_espanol}.xlsx')
ws = wb.active

for cell in ws['F'][1:]:
    cell.number_format = '$#,##0.00'
    wb.save(f'{mes_actual_espanol}.xlsx')

wb = openpyxl.load_workbook(f'{mes_actual_espanol}.xlsx')
ws = wb.active

for cell in ws['G'][1:]:
    cell.number_format = '$#,##0.00'
    wb.save(f'{mes_actual_espanol}.xlsx')

wb = openpyxl.load_workbook(f'{mes_actual_espanol}.xlsx')
ws = wb.active

for cell in ws['H'][1:]:
    cell.number_format = '$#,##0.00'
# Guardar los cambios
wb.save(f'{mes_actual_espanol}.xlsx')